#! /usr/bin/env python
# _*_ coding: utf-8 _*_

# 外部库（需要“pip install 库名”进行安装）
from openpyxl import load_workbook
import arrow

# 内置库
import re
import os

# 自定义模块
import fetchs

# 精准四舍五入保留 1 位小数
def round_decimal(value):
    m = re.search(r'\.(\d)5$', str(value))
    if m and int(m.group(1)) % 2 == 0:
        return round(value, 1) + 0.1
    else:
        return round(value * 1.0, 1)


# 计算变化幅度，保留 1 位小数
def calc_relative_ratio_1(v1, v2):
    if v1 == 0 and v2 == 0:
        return '-'
    elif v1 == 0:
        return '+∞'
    else:
        return str(round_decimal((v2 - v1) / v1 * 100)) + '%'


# 计算变化幅度，保留 2 位小数
def calc_relative_ratio_2(v1, v2):
    if v1 == 0 and v2 == 0:
        return '-'
    elif v1 == 0:
        return '+∞'
    else:
        return str(round((v2 - v1) / v1 * 100, 2)) + '%'


# 获取字段名和列索引的映射
def get_fields(row):
    fields = {}
    for i, e in enumerate(row):
        if e.value:
            fields[e.value.strip()] = i
        else:
            continue
    return fields


def get_articles(filepath):
    wb = load_workbook(filepath, read_only=True)

    if len(wb.sheetnames) > 1:
        print('以下表格存在多个数据表，除第一个以外的都要删除：')
        print(filepath)
        exit()

    sheet_name = wb.sheetnames[0]
    if not (sheet_name == '数据清单' or sheet_name == 'list'):
        print('源数据表表名不正确（应为“数据清单”或“list”），无法读取，程序退出')
        exit()
    
    ws = wb.get_sheet_by_name(sheet_name)
    fields = get_fields(ws[1])
    articles = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        article = {}
        if sheet_name == '数据清单': # 微博
            article['id'] = row[fields['主页链接']][-10:]
            article['name'] = row[fields['微博昵称']]
            article['follower'] = row[fields['粉丝数']]
            article['publish_time'] = row[fields['发布时间']]
            if isinstance(article['publish_time'], str):
                article['publish_time'] = arrow.get(article['publish_time']).datetime
            article['content'] = row[fields['内容']]
            article['quote'] = row[fields['转发数']]
            article['comment'] = row[fields['评论数']]
            article['like'] = row[fields['点赞数']]
            article['url'] = row[fields['链接']]
            article['homepage_url'] = row[fields['主页链接']]
        else: # 微信
            article['name'] = row[fields['公众号昵称']] if '公众号昵称' in fields else row[fields['账号名称']]
            article['id'] = row[fields['微信号']]
            if '@qianyi' in article['id']:
                article['id'] = re.sub(r'@qianyi', '', article['id'])
            article['author'] = row[fields['作者']]
            article['is_head'] = '是' if row[fields['发布位置']] == 0 else '否'
            article['is_original'] = '是' if row[fields['是否原创']] == 1 else '否'
            article['video_count'] = row[fields['含视频数量']]
            article['title'] = row[fields['标题']]
            article['url'] = row[fields['文章链接']]
            article['abstract'] = row[fields['摘要']]
            article['content'] = row[fields['正文']]
            article['read'] = row[fields['阅读数']]
            article['like'] = row[fields['点赞数']]
            article['reward'] = row[fields['赞赏数']]
            article['comment'] = row[fields['评论数']]
            article['comment_like'] = row[fields['评论总点赞数']]
            article['comment_reply'] = row[fields['回复评论数']]
            article['comment_reply_like'] = row[fields['回复评论总点赞数']]
            article['image_url'] = row[fields['图片链接']]
            article['origin_url'] = row[fields['原文链接']]
            article['video_url'] = row[fields['视频链接']]
            article['music_url'] = row[fields['音乐链接']]
            article['audio_url'] = row[fields['音频链接']]
            article['memo'] = row[fields['备注']]
            article['publish_time'] = row[fields['发布时间']]
            if isinstance(article['publish_time'], str):
                article['publish_time'] = arrow.get(article['publish_time']).datetime
        articles.append(article)
    return articles


def get_followers(filepath):
    wb = load_workbook(filepath, read_only=True)
    sheet_name = wb.sheetnames[0]

    ws = wb.get_sheet_by_name(sheet_name)
    fields = get_fields(ws[1])
    followers = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        follower = {}
        follower['oid'] = row[fields['oid']]
        follower['author'] = row[fields['昵称']]
        follower['gender'] = row[fields['性别']]
        follower['province'] = row[fields['省份']]
        follower['city'] = row[fields['城市']]
        follower['home_url'] = row[fields['主页链接']]
        follower['from'] = str(row[fields['来自']])
        followers.append(follower)
    return followers


def get_mentions(filepath):
    wb = load_workbook(filepath, read_only=True)
    sheet_name = wb.sheetnames[0]

    ws = wb.get_sheet_by_name(sheet_name)
    fields = get_fields(ws[1])
    mentions = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        mention = {}
        mention['keyword'] = row[fields['关键词']]
        mention['quote_count'] = row[fields['转发数']]
        mention['comment_count'] = row[fields['评论数']]
        mention['author'] = row[fields['作者昵称']]
        mention['oid'] = str(row[fields['uid']])
        mention['follower_count'] = row[fields['粉丝数']]
        mention['url'] = row[fields['链接']]
        mention['content'] = row[fields['内容']]
        mentions.append(mention)
    return mentions


def get_ambiguous_words(brand):
    filepath = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'src', 'ambiguous_words_' + brand + '.txt')
    with open(filepath, 'r', encoding='utf-8') as f:
        words = list(map(lambda i: i[:-1] if '\n' in i else i, f.readlines()))
    return words


def group_weibo_by_brand(articles):
    groups = []
    for article in articles:
        is_new = True
        for group in groups:
            if group['brand'] == article['brand']:
                is_new = False
                group['article_count'] += 1
                group['quote_sum'] += article['quote']
                group['comment_sum'] += article['comment']
                group['like_sum'] += article['like']
                group['interact_sum'] = group['interact_sum'] + article['quote'] + article['comment'] + article['like']

        if is_new:
            groups.append({
                'brand': article['brand'],
                'article_count': 1,
                'quote_sum': article['quote'],
                'comment_sum': article['comment'],
                'like_sum': article['like'],
                'interact_sum': article['quote'] + article['comment'] + article['like']
            })
    
    return groups


def group_weixin_by_brand(articles):
    groups = []
    for article in articles:
        is_new = True
        for group in groups:
            if group['brand'] == article['brand']:
                is_new = False
                group['article_count'] += 1
                group['read_sum'] += article['read']
                group['like_sum'] += article['like']
                group['comment_sum'] += article['comment']
                group['comment_like_sum'] += article['comment_like']
        if is_new:
            groups.append({
                'brand': article['brand'],
                'article_count': 1,
                'read_sum': article['read'],
                'like_sum': article['like'],
                'comment_sum': article['comment'],
                'comment_like_sum': article['comment_like']
            })
    
    return groups


def get_videos(filepath):
    wb = load_workbook(filepath, read_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb.get_sheet_by_name(sheet_name)

    fields= get_fields(ws[1])
    source_type_index = fields['source_type']
    url_index = fields['url']
    publish_time_index = fields['release_date']
    title_index = fields['title']
    videos = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[source_type_index] == '6':
            video = {
                'brand': match_brand(row[title_index]),
                'url': row[url_index],
                'publish_time': str(row[publish_time_index]),
                'title': row[title_index],
                'platform': match_platform(row[url_index]),
            }
            videos.append(video)
        else:
            continue
    
    for video in videos:
        if video['platform'] == '腾讯视频':
            r = fetchs.qqvideo(video['url'])
            video['id'] = r['video_id']
            video['impress'] = r['play_num']
        elif video['platform'] == '爱奇艺':
            r = fetchs.iqiyi(video['url'])
            video['id'] = r['video_id']
            video['impress'] = r['play_num']
        elif video['platform'] == '芒果TV':
            r = fetchs.mgtv(video['url'])
            video['id'] = r['video_id']
            video['impress'] = r['play_num']
            
        else:
            video['id'] = None
            video['impress'] = None
    
    return videos


def get_datas(filepath):
    wb = load_workbook(filepath, read_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb.get_sheet_by_name(sheet_name)

    fields= get_fields(ws[1])
    url_crc_index = fields['url_crc']
    url_index = fields['url']
    title_index = fields['title']
    source_type_index = fields['source_type']
    release_date_index = fields['release_date']
    datas = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        data = {
            'url_crc': row[url_crc_index],
            'url': row[url_index],
            'title': row[title_index],
            'source_type': row[source_type_index],
            'release_date': row[release_date_index]}
        datas.append(data)
    return datas


def match_brand(str):
    if re.search(r'evcard', str, re.I) or str == 'E享会':
        return 'EVCARD'
    elif re.search(r'gofun', str, re.I):
        return 'GoFun'
    elif '盼达' in str:
        return '盼达用车'
    elif re.search(r'car2go', str, re.I):
        return 'car2go'
    elif '途歌' in str:
        return '途歌'
    elif '摩范出行' in str:
        return '摩范出行'
    elif re.search(r'ponycar', str, re.I):
        return 'PonyCar'
    elif '小桔租车' in str or '滴滴共享汽车' in str:
        return '小桔租车'
    else:
        return '其他'


def match_platform(str):
    if 'qq.com' in str:
        return '腾讯视频'
    elif 'iqiyi' in str:
        return '爱奇艺'
    elif 'mgtv' in str:
        return '芒果TV'
    else:
        return '其他'