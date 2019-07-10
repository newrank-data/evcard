#! /usr/bin/env python
# _*_ coding: utf-8 _*_

# 此脚本用于整理双微回采的数据，生成最终提交给客户的双微数据表，包含 1 个汇总表、1 个微博数据表和 1 个微信数据表

# 外部库（需要“pip install 库名”进行安装）
import arrow
from pymongo import MongoClient
from openpyxl import Workbook

# 内置库
import os
from bson.objectid import ObjectId

# 自定义模块
import utils

# 数据库配置
client = MongoClient()
db = client.newrank
weibo_collection = db.sharecar_weibo_account
weixin_collection = db.sharecar_weixin_account

# 初始化
label = arrow.now().shift(months=-1).format('YYYYMM')
weibo_filename = 'weibo_' + label + '.xlsx'
weibo_filepath = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'src', weibo_filename)
weixin_filename = 'weixin_' + label + '.xlsx'
weixin_filepath = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'src', weixin_filename)
brands = {
    'evcard': 'EVCARD',
    'gofun': 'GoFun',
    'panda': '盼达用车',
    'car2go': 'car2go',
    'togo': '途歌',
    'morefun': '摩范出行',
    'ponycar': 'PonyCar',
    'xiaoju': '小桔租车'
    }

primary_weibo_accounts = {
    'EVCARD': 'EVCARD官方微博',
    'GoFun': 'GoFun出行',
    '盼达用车': '盼达用车',
    'car2go': '即行car2go',
    '途歌': 'TOGO途歌',
    '摩范出行': '摩范出行',
    'PonyCar': 'PonyCar马上用车',
    '小桔租车': '小桔租车官方微博'
}

primary_weixin_accounts = {
    'EVCARD': 'EVCARD服务号',
    'GoFun': 'GoFun出行',
    '盼达用车': '盼达用车',
    'car2go': '即行car2go',
    '途歌': 'TOGO途歌',
    '摩范出行': '摩范出行',
    'PonyCar': 'PONYCAR马上用车',
    '小桔租车': '小桔租车平台'
}
# 微博匹配账号信息
weibo_articles = utils.get_articles(weibo_filepath)
weibo_articles.sort(key=lambda article: article['id'])
iter_weibo_account = None
for article in weibo_articles:

    # 当 id 变化时，更新用于匹配的迭代账号
    if not (iter_weibo_account and iter_weibo_account['id'] == article['id']):
        iter_weibo_account = weibo_collection.find_one(
            {'id': article['id']},
            {'_id': 0, 'id': 1, 'brand': 1, 'is_primary': 1, 'is_regional': 1, 'region': 1}
        )

    article['brand'] = brands.get(iter_weibo_account['brand'])
    article['is_primary'] = '是' if iter_weibo_account['is_primary'] == 1 else '否'
    article['region'] = iter_weibo_account.get('region', '')

weibo_articles.sort(key=lambda article: article['brand'])


# 微信匹配账号信息
weixin_articles = utils.get_articles(weixin_filepath)
weixin_articles.sort(key=lambda article: article['id'])
iter_weixin_account = None
for article in weixin_articles:

    # 当 id 变化时，更新用于匹配的迭代账号
    if not (iter_weixin_account and iter_weixin_account['id'] == article['id']):
        iter_weixin_account = weixin_collection.find_one(
            {'id': article['id']},
            {'_id': 0, 'id': 1, 'brand': 1, 'type': 1, 'is_primary': 1, 'is_regional': 1, 'region': 1}
        )

    if not iter_weixin_account:
        print('无法在数据库中匹配到公众号：', article['id'])
        exit()
    
    article['brand'] = brands.get(iter_weixin_account['brand'])
    article['type'] = '订阅号' if iter_weixin_account['type'] == 1 else '服务号'
    article['is_primary'] = '是' if iter_weixin_account['is_primary'] == 1 else '否'
    article['region'] = iter_weixin_account.get('region', '')
    
weixin_articles.sort(key=lambda article: article['brand'])


# 创建工作簿，分别写入微博和微信工作表，存到 dist 目录
wb = Workbook()

# 表1【汇总】
ws1 = wb.active
ws1.title = '汇总'

# 所有微博
ws1.append(['EVCARD 及竞品微博所有帐号数据总览'])
ws1.append(['品牌', '微博数', '转发数', '评论数', '点赞数', '互动总数'])
weibo_brand_groups = utils.group_weibo_by_brand(weibo_articles)
for group in weibo_brand_groups:
    ws1.append([group['brand'], group['article_count'], group['quote_sum'], group['comment_sum'], group['like_sum'], group['interact_sum']])

# 微博主号
ws1.append([''])
ws1.append(['EVCARD 及竞品微博主号数据总览'])
ws1.append(['微博昵称', '微博数', '转发数', '评论数', '点赞数', '互动总数'])
primary_weibo_brand_groups = utils.group_weibo_by_brand(list(filter(lambda article: article['is_primary'] == '是', weibo_articles)))
for group in primary_weibo_brand_groups:
    ws1.append([primary_weibo_accounts[group['brand']], group['article_count'], group['quote_sum'],
    group['comment_sum'], group['like_sum'], group['interact_sum']])

# 所有微信
ws1.append([''])
ws1.append(['EVCARD 及竞品微信所有帐号数据总览'])
ws1.append(['品牌', '发文数', '阅读数', '点赞数', '评论数', '评论总点赞数'])
weixin_brand_groups = utils.group_weixin_by_brand(weixin_articles)
for group in weixin_brand_groups:
    ws1.append([group['brand'], group['article_count'], group['read_sum'], group['like_sum'], group['comment_sum'], group['comment_like_sum']])

# 微信服务号
ws1.append([''])
ws1.append(['EVCARD 及竞品微信服务号数据总览'])
ws1.append(['账号名称', '发文数', '阅读数', '点赞数', '评论数', '评论总点赞数'])
type2_weixin_brand_groups = utils.group_weixin_by_brand(list(filter(lambda article: article['type'] == '服务号', weixin_articles)))
for group in type2_weixin_brand_groups:
    ws1.append([primary_weixin_accounts[group['brand']], group['article_count'], group['read_sum'],
    group['like_sum'], group['comment_sum'], group['comment_like_sum']])

# 微信订阅号
ws1.append([''])
ws1.append(['EVCARD 及竞品微信所有订阅号数据总览'])
ws1.append(['品牌', '发文数', '阅读数', '点赞数', '评论数', '评论总点赞数'])
type1_weixin_brand_groups = utils.group_weixin_by_brand(list(filter(lambda article: article['type'] == '订阅号', weixin_articles)))
for group in type1_weixin_brand_groups:
    ws1.append([group['brand'], group['article_count'], group['read_sum'], group['like_sum'], group['comment_sum'], group['comment_like_sum']])

# 微信主号
ws1.append([''])
ws1.append(['EVCARD 及竞品微信主号数据总览'])
ws1.append(['账号名称', '发文数', '阅读数', '点赞数', '评论数', '评论总点赞数'])
primary_weixin_brand_groups = utils.group_weixin_by_brand(list(filter(lambda article: article['is_primary'] == '是', weixin_articles)))
for group in primary_weixin_brand_groups:
    ws1.append([primary_weixin_accounts[group['brand']], group['article_count'], group['read_sum'], group['like_sum'], group['comment_sum'], group['comment_like_sum']])

# 表2【微博】
ws2 = wb.create_sheet(title='微博')
for index, article in enumerate(weibo_articles):
    if index == 0:
        ws2.append(['品牌', '是否主号', '地区', '微博昵称', '粉丝数', '发布时间', '内容', '转发数', '评论数', '点赞数', '链接', '主页链接'])
    else:
        ws2.append([
            article['brand'], article['is_primary'], article['region'], article['name'], int(article['follower']), str(article['publish_time']),
            article['content'], article['quote'], article['comment'], article['like'], article['url'], article['homepage_url']
        ])

# 表3【微信】
ws3 = wb.create_sheet(title='微信')
for index, article in enumerate(weixin_articles):
    if index == 0:
        ws3.append([
            '品牌', '账号类型', '是否主号', '地区', '公众号昵称', '微信号', '作者',
            '是否头条', '是否原创', '含视频数量', '标题', '文章链接', '摘要', '正文',
            '阅读数', '点赞数', '赞赏数', '评论数', '评论总点赞数', '回复评论数', '回复评论总点赞数',
            '图片链接', '原文链接', '视频链接', '音乐链接', '音频链接', '备注', '发布时间'
        ])
    else:
        ws3.append([
            article['brand'], article['type'], article['is_primary'], article['region'],
            article['name'], article['id'], article['author'], article['is_head'],
            article['is_original'], article['video_count'], article['title'], article['url'],
            article['abstract'], article['content'], article['read'], article['like'],
            article['reward'], article['comment'], article['comment_like'], article['comment_reply'],
            article['comment_reply_like'], article['image_url'], article['origin_url'], article['video_url'],
            article['music_url'], article['audio_url'], article['memo'], article['publish_time'],
        ])

wb.save('./dist/双微数据表_{}.xlsx'.format(label))
print('完成！结果文件保存在 dist 目录下')